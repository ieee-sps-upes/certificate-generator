"""
import_from_excel.py
Imports all participants from 'final table.xlsx' into the ieee_participants collection.
Run from the backend/ directory: python import_from_excel.py
"""
import io, sys
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
import sys
import os

# Allow running from anywhere by resolving the xlsx path
EXCEL_PATH = os.path.join(os.path.dirname(__file__), '..', 'final table.xlsx')

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Installing...")
    os.system(f"{sys.executable} -m pip install openpyxl")
    import openpyxl

from database import get_db

def import_participants():
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active
    headers = [cell.value for cell in ws[1]]
    rows = [dict(zip(headers, [cell.value for cell in row])) for row in ws.iter_rows(min_row=2)]

    db = get_db()
    col = db["ieee_participants"]

    inserted, updated, skipped = 0, 0, 0

    for row in rows:
        name  = (str(row.get('Name') or '')).strip()
        sap   = str(row.get('SAP ID') or '').strip()
        email = (str(row.get('Email') or '')).strip() or None

        if not name or not sap:
            print(f"  [!] Skipping incomplete row: {row}")
            skipped += 1
            continue

        sap_str = sap  # store as string

        # Check for existing record by SAP ID
        existing = col.find_one({"sap_id": sap_str})

        doc = {"name": name, "sap_id": sap_str}
        if email:
            doc["email"] = email
        # If no email, we do NOT store an email — the user will provide one at claim time

        if existing:
            col.update_one({"sap_id": sap_str}, {"$set": doc})
            updated += 1
            status = f"email={email}" if email else "no email"
            print(f"  [U] Updated  : {name} ({status})")
        else:
            col.insert_one(doc)
            inserted += 1
            status = f"email={email}" if email else "no email"
            print(f"  [+] Inserted : {name} ({status})")

    print(f"\n=== Done ===")
    print(f"  Inserted : {inserted}")
    print(f"  Updated  : {updated}")
    print(f"  Skipped  : {skipped}")
    print(f"  Total    : {len(rows)}")


if __name__ == "__main__":
    print("=== IEEE Participant Importer ===")
    print(f"Reading from: {os.path.abspath(EXCEL_PATH)}\n")
    try:
        import_participants()
    except Exception as e:
        import traceback
        print(f"Error: {e}")
        print(traceback.format_exc())
        print("Make sure MongoDB is running and the Excel file is present.")
